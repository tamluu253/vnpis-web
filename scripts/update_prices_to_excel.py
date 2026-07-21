import pdfplumber
import pandas as pd
from openpyxl import load_workbook

pdf1 = r"C:\Users\TL\Documents\Disk D\VNPIS\###Suppliers\Sanjin\Price list for Agent （20260717）.pdf"
pdf2 = r"C:\Users\TL\Documents\Disk D\VNPIS\###Suppliers\Sanjin\Price list for Agent （20260717） - sreen printer and hot stamping.pdf"
excel_path = r"C:\Users\TL\Documents\Disk D\VNPIS\###Suppliers\Sanjin\BaoGia_MayInSJ_NoiBo.xlsx"

# 1. Extract prices from PDFs
prices = {} # model -> price
prefix_prices = {} # prefix -> price

def extract_prices_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if len(row) >= 7:
                        prefix = str(row[0]).strip().upper()
                        model = str(row[2]).strip().upper()
                        price = str(row[6]).strip()
                        
                        if price and price.isdigit():
                            if model:
                                prices[model] = int(price)
                            if prefix:
                                prefix_prices[prefix] = int(price)

extract_prices_from_pdf(pdf1)
extract_prices_from_pdf(pdf2)

print(f"Extracted {len(prices)} prices from PDFs.")

# 2. Update Excel file
wb = load_workbook(excel_path)
ws = wb['Báo Giá Máy SJ']

# Find column indices
headers = {cell.value: i for i, cell in enumerate(ws[1])}
model_col = headers.get('Model')
price_col = headers.get('Giá mua (Tệ/VNĐ)')
ref_col = headers.get('File Tham Khảo (PDF)')

if model_col is None or price_col is None:
    print("Error: Could not find columns in Excel.")
    exit(1)

updated_count = 0

for row in ws.iter_rows(min_row=2):
    model_cell = row[model_col]
    price_cell = row[price_col]
    ref_cell = row[ref_col] if ref_col is not None else None
    
    if not model_cell.value:
        continue
        
    model = str(model_cell.value).strip().upper()
    price = None
    
    # 1. Try exact model match
    if model in prices:
        price = prices[model]
    else:
        # 2. Try prefix match (from reference file if we extracted it)
        # For example, ref_cell.value might be "A5(SYDK-125-100).pdf", we can extract "A5"
        if ref_cell and ref_cell.value:
            ref_val = str(ref_cell.value)
            # Find prefix like A1, B10, etc at start of ref file
            import re
            match = re.match(r'^([A-Z]\d+)', ref_val.upper())
            if match:
                prefix = match.group(1)
                if prefix in prefix_prices:
                    price = prefix_prices[prefix]
    
    if price is not None:
        price_cell.value = price
        updated_count += 1

wb.save(excel_path)
print(f"Updated {updated_count} rows with prices in {excel_path}")
